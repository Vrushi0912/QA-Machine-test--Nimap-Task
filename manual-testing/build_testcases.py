import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
MODULE_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
BODY_FONT = Font(name="Arial", size=10)
WRAP = Alignment(wrap_text=True, vertical="top")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def style_header(ws, row=1):
    for cell in ws[row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER

def style_body(ws, start_row, end_row, end_col):
    for r in range(start_row, end_row + 1):
        for c in range(1, end_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = BODY_FONT
            cell.alignment = WRAP
            cell.border = BORDER

def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

# ============================================================
# SHEET 1: TEST CASES (all 4 modules)
# ============================================================
ws = wb.active
ws.title = "Test Cases"
headers = ["Module", "TC ID", "Test Scenario", "Test Steps", "Test Data",
           "Expected Result", "Actual Result", "Status (Pass/Fail/Blocked)", "Priority", "Bug ID (if failed)"]
ws.append(headers)
style_header(ws)
ws.freeze_panes = "A2"

rows = []

# ---------------- SIGN UP ----------------
m = "Sign Up"
rows += [
 [m,"TC_SU_01","Verify Sign Up page loads with all mandatory fields","1. Navigate to app URL\n2. Click 'Sign Up'","N/A","Sign Up form loads with Name, Email, Mobile No., Password, Confirm Password fields and Submit button","","","High",""],
 [m,"TC_SU_02","Verify successful signup with valid, unique details","1. Enter valid unique Name, Email, Mobile, Password, Confirm Password\n2. Click Submit","Name: Test User\nEmail: newuser01@test.com\nMobile: 9876543210\nPassword: Test@1234","User is registered successfully; confirmation toast/message shown; user redirected to Login or OTP screen","","","High",""],
 [m,"TC_SU_03","Verify signup fails with an already registered email","1. Enter an email already used to sign up\n2. Fill remaining valid fields\n3. Click Submit","Email: existinguser@test.com","Error message 'Email already registered' is displayed; form not submitted","","","High",""],
 [m,"TC_SU_04","Verify validation when mandatory field(s) left blank","1. Leave Name field blank\n2. Fill remaining fields\n3. Click Submit","Name: (blank)","Inline validation message 'Name is required' shown; form not submitted","","","High",""],
 [m,"TC_SU_05","Verify email field rejects invalid email format","1. Enter invalid email format\n2. Click Submit / Tab out","Email: user@@test, user.com, user@.com","Validation message 'Enter a valid email address' displayed","","","Medium",""],
 [m,"TC_SU_06","Verify mobile number field rejects invalid input","1. Enter mobile number with letters/less than required digits\n2. Click Submit","Mobile: abcd123456, 98765","Validation message 'Enter a valid 10-digit mobile number' displayed","","","Medium",""],
 [m,"TC_SU_07","Verify password and confirm password mismatch is caught","1. Enter Password\n2. Enter a different Confirm Password\n3. Click Submit","Password: Test@1234\nConfirm Password: Test@5678","Error 'Passwords do not match' displayed; form not submitted","","","High",""],
 [m,"TC_SU_08","Verify password field enforces minimum complexity","1. Enter a weak password (e.g., only lowercase, <6 chars)\n2. Click Submit","Password: abc","Validation message about password policy (length/uppercase/number/special char) displayed","","","Medium",""],
 [m,"TC_SU_09","Verify password field masks input by default","1. Enter password\n2. Observe field","Password: Test@1234","Characters shown as dots/asterisks; optional 'show password' eye icon toggles visibility","","","Low",""],
 [m,"TC_SU_10","Verify leading/trailing spaces are trimmed or rejected in Name/Email","1. Enter '  Test User  ' in Name field\n2. Submit","Name: '  Test User  '","Either spaces are trimmed automatically or a validation message is shown (define expected behavior with dev team)","","","Low",""],
 [m,"TC_SU_11","Verify SQL Injection / script tags are handled safely in text fields","1. Enter script/SQL payload in Name field\n2. Submit","Name: <script>alert(1)</script>","Input is sanitized/rejected; no script executes; no application error/crash","","","High",""],
 [m,"TC_SU_12","Verify Terms & Conditions checkbox (if present) is mandatory","1. Leave T&C checkbox unchecked\n2. Fill rest of form\n3. Click Submit","N/A","Error 'Please accept Terms & Conditions' shown; form not submitted","","","Medium",""],
 [m,"TC_SU_13","Verify Sign Up button is disabled/no-op while mandatory fields are empty","1. Load Sign Up page\n2. Without entering data, click Submit","N/A","Form does not submit; validation errors highlighted on empty required fields","","","Medium",""],
 [m,"TC_SU_14","Verify navigation link back to Login page from Sign Up screen","1. On Sign Up page, click 'Already have an account? Login'","N/A","User is navigated to the Login screen","","","Low",""],
]

# ---------------- FORGOT PASSWORD ----------------
m = "Forgot Password"
rows += [
 [m,"TC_FP_01","Verify 'Forgot Password' link is visible and clickable on Login page","1. Navigate to Login page\n2. Locate 'Forgot Password' link","N/A","Link is visible, clicking navigates to Forgot Password screen","","","Medium",""],
 [m,"TC_FP_02","Verify Forgot Password with a valid registered email/mobile","1. Enter registered email/mobile\n2. Click Submit/Send OTP","Email: registereduser@test.com","Success message shown; reset link/OTP sent to registered email/mobile","","","High",""],
 [m,"TC_FP_03","Verify Forgot Password with an unregistered email/mobile","1. Enter an email/mobile not registered in the system\n2. Click Submit","Email: notregistered@test.com","Error message 'No account found with this email' displayed (message should not reveal whether it's email or mobile that's wrong, for security)","","","High",""],
 [m,"TC_FP_04","Verify Forgot Password with blank email/mobile field","1. Leave field blank\n2. Click Submit","N/A","Validation message 'This field is required' displayed","","","Medium",""],
 [m,"TC_FP_05","Verify Forgot Password with invalid email format","1. Enter malformed email\n2. Click Submit","Email: user@@test","Validation message 'Enter a valid email' displayed","","","Medium",""],
 [m,"TC_FP_06","Verify reset link/OTP expires after the defined validity window","1. Request reset link/OTP\n2. Wait beyond expiry time (e.g., 10/15 min)\n3. Try to use it","N/A","Expired link/OTP is rejected with an appropriate 'link/OTP expired' message","","","High",""],
 [m,"TC_FP_07","Verify user can set a new password meeting policy after clicking reset link","1. Click valid reset link\n2. Enter new Password + Confirm Password meeting policy\n3. Submit","New Password: NewTest@123","Password reset successful; confirmation shown; user can log in with new password","","","High",""],
 [m,"TC_FP_08","Verify new password cannot be same as old password (if business rule exists)","1. Attempt to reset password to the same current password","Password: (same as current)","Error 'New password must be different from old password' displayed (only if this rule applies)","","","Medium",""],
 [m,"TC_FP_09","Verify New Password and Confirm Password mismatch is caught","1. Enter different values in New Password and Confirm Password\n2. Submit","New: Test@123 / Confirm: Test@456","Error 'Passwords do not match' displayed","","","High",""],
 [m,"TC_FP_10","Verify a used/already-consumed reset link cannot be reused","1. Use a reset link successfully once\n2. Try to reuse the same link","N/A","Error message indicating the link is no longer valid","","","Medium",""],
 [m,"TC_FP_11","Verify multiple rapid Forgot Password requests are rate-limited","1. Submit Forgot Password request multiple times in quick succession","N/A","System throttles/blocks after N attempts with an appropriate message (prevents abuse/spam)","","","Medium",""],
]

# ---------------- SIGN WITH OTP ----------------
m = "Sign with OTP"
rows += [
 [m,"TC_OTP_01","Verify 'Login/Sign with OTP' option is available and navigable","1. On Login page, select 'Login with OTP' option","N/A","User is navigated to mobile/email entry screen for OTP login","","","Medium",""],
 [m,"TC_OTP_02","Verify OTP is sent to a valid registered mobile number/email","1. Enter registered mobile/email\n2. Click 'Send OTP'","Mobile: 9876543210 (registered)","OTP sent; confirmation message 'OTP sent to XXXXXX210' shown; OTP input field appears","","","High",""],
 [m,"TC_OTP_03","Verify OTP request fails for unregistered mobile/email","1. Enter unregistered mobile/email\n2. Click 'Send OTP'","Mobile: 9999999999 (not registered)","Error message 'No account found' or similar displayed; OTP not sent","","","High",""],
 [m,"TC_OTP_04","Verify successful login with correct OTP","1. Request OTP\n2. Enter correct OTP received\n3. Click Verify/Submit","OTP: <valid 4/6-digit code>","User logged in successfully and redirected to dashboard","","","High",""],
 [m,"TC_OTP_05","Verify login fails with incorrect OTP","1. Request OTP\n2. Enter an incorrect OTP\n3. Click Verify","OTP: 000000 (wrong)","Error message 'Invalid OTP' displayed; user remains on OTP screen","","","High",""],
 [m,"TC_OTP_06","Verify OTP expires after defined validity period","1. Request OTP\n2. Wait beyond expiry (e.g., 2/5 min)\n3. Enter the expired OTP","N/A","Error message 'OTP expired, please resend' displayed","","","High",""],
 [m,"TC_OTP_07","Verify 'Resend OTP' functionality and its cooldown timer","1. Request OTP\n2. Click 'Resend OTP' immediately\n3. Click 'Resend OTP' after cooldown expires","N/A","Resend is disabled/greyed out during cooldown (e.g., 30s countdown shown); works after cooldown ends and a new OTP is sent","","","Medium",""],
 [m,"TC_OTP_08","Verify OTP field only accepts numeric input of correct length","1. Attempt to enter letters/special characters in OTP field\n2. Attempt to enter more/fewer digits than expected","OTP: abcd, 12","Non-numeric input rejected; field restricts to expected digit length","","","Medium",""],
 [m,"TC_OTP_09","Verify account lockout/throttling after multiple wrong OTP attempts","1. Enter wrong OTP repeatedly (e.g., 5 times)","OTP: wrong x5","Account/OTP request is temporarily locked with a warning message, preventing brute-force attempts","","","High",""],
 [m,"TC_OTP_10","Verify OTP screen allows navigating back to change mobile/email","1. On OTP entry screen, click 'Change Number/Email' or back option","N/A","User is returned to number/email entry screen to correct and resend","","","Low",""],
]

# ---------------- LOGIN ----------------
m = "Login"
rows += [
 [m,"TC_LG_01","Verify Login page loads with Username/Email and Password fields","1. Navigate to app URL","N/A","Login form displayed with Email/Username, Password fields, Login button, Forgot Password & Sign Up links","","","High",""],
 [m,"TC_LG_02","Verify successful login with valid registered credentials","1. Enter valid Email/Username and Password\n2. Click Login","Email: registereduser@test.com\nPassword: Test@1234","User logged in successfully; redirected to Dashboard","","","High",""],
 [m,"TC_LG_03","Verify login fails with invalid password","1. Enter valid Email\n2. Enter incorrect Password\n3. Click Login","Email: registereduser@test.com\nPassword: WrongPass","Error message 'Invalid credentials / Incorrect password' displayed; user stays on Login page","","","High",""],
 [m,"TC_LG_04","Verify login fails with unregistered email/username","1. Enter an email not registered\n2. Enter any password\n3. Click Login","Email: notexist@test.com","Error message 'Account not found / Invalid credentials' displayed","","","High",""],
 [m,"TC_LG_05","Verify validation when Email/Password left blank","1. Leave Email blank, click Login\n2. Leave Password blank, click Login","N/A","Respective inline validation message 'This field is required' shown for each","","","Medium",""],
 [m,"TC_LG_06","Verify Password field masks the entered password","1. Enter password and observe field","Password: Test@1234","Password shown as dots/asterisks; optional show/hide toggle works correctly","","","Low",""],
 [m,"TC_LG_07","Verify 'Remember Me' checkbox (if present) persists session/credentials correctly","1. Check 'Remember Me'\n2. Login successfully\n3. Close and reopen browser","N/A","User session/credentials retained per expected behavior when checkbox is checked","","","Low",""],
 [m,"TC_LG_08","Verify account lockout after multiple consecutive failed login attempts","1. Attempt login with wrong password 5 times","Password: wrong x5","Account is temporarily locked / CAPTCHA triggered with an appropriate warning message","","","High",""],
 [m,"TC_LG_09","Verify SQL Injection payload in login fields is handled safely","1. Enter SQL injection string in Email/Password field\n2. Click Login","Email: ' OR '1'='1","Login is rejected safely; no unauthorized access granted; no application error exposed","","","High",""],
 [m,"TC_LG_10","Verify session expires appropriately after prolonged inactivity","1. Login successfully\n2. Leave the session idle beyond timeout period\n3. Try to perform an action","N/A","User is logged out / redirected to login page with a session-expired message","","","Medium",""],
 [m,"TC_LG_11","Verify user cannot access dashboard URL directly without logging in","1. Log out (or don't log in)\n2. Directly hit the dashboard/internal URL","N/A","User is redirected to Login page; internal page not accessible without authentication","","","High",""],
 [m,"TC_LG_12","Verify login page is responsive across screen sizes/browsers","1. Open Login page on Chrome, Firefox, Edge and on a mobile viewport","N/A","Layout renders correctly, fields and buttons are usable on all tested browsers/screen sizes","","","Low",""],
]

for r in rows:
    ws.append(r)

last_row = ws.max_row
style_body(ws, 2, last_row, len(headers))

# Shade module column groups + module separator styling
for r in range(2, last_row + 1):
    ws.cell(row=r, column=1).font = Font(name="Arial", bold=True, size=10)

set_widths(ws, [14, 10, 30, 30, 22, 32, 16, 16, 10, 14])
for r in range(2, last_row + 1):
    ws.row_dimensions[r].height = 45

# ============================================================
# SHEET 2: FIELD VALIDATIONS
# ============================================================
ws2 = wb.create_sheet("Field Validations")
headers2 = ["Module", "Field Name", "Field Type", "Validation Rule", "Valid Input Example",
            "Invalid Input Example", "Expected System Behavior"]
ws2.append(headers2)
style_header(ws2)
ws2.freeze_panes = "A2"

fv_rows = [
 ["Sign Up","Name","Text","Required; alphabets & spaces only; min 2, max 50 chars","Test User","Te$t123, (blank)","Reject numbers/special chars; show 'Invalid name' / 'Required' message"],
 ["Sign Up","Email","Email","Required; standard email regex; must be unique","user@test.com","user@@test, user.com","Show 'Invalid email format' or 'Email already registered'"],
 ["Sign Up","Mobile Number","Numeric","Required; exactly 10 digits; numeric only; unique","9876543210","98765, abc1234567","Show 'Enter valid 10-digit mobile number'"],
 ["Sign Up","Password","Password","Required; min 8 chars; 1 upper, 1 lower, 1 digit, 1 special char","Test@1234","test, 12345678","Show password policy message; mask input"],
 ["Sign Up","Confirm Password","Password","Required; must exactly match Password field","Test@1234","Test@5678","Show 'Passwords do not match'"],
 ["Forgot Password","Registered Email/Mobile","Email/Numeric","Required; must exist in system","registered@test.com","notexist@test.com","Show 'No account found' for unregistered value"],
 ["Forgot Password","New Password","Password","Required; same policy as Sign Up password; must differ from old (if rule exists)","NewTest@123","abc","Show policy error / 'must differ from old password'"],
 ["Forgot Password","Confirm New Password","Password","Required; must match New Password","NewTest@123","NewTest@999","Show 'Passwords do not match'"],
 ["Sign with OTP","Mobile Number/Email","Numeric/Email","Required; must be registered; valid format","9876543210","invalidmobile","Show format/registration error"],
 ["Sign with OTP","OTP","Numeric","Required; fixed length (4 or 6 digits); numeric only; time-bound validity","123456","12ab, 000000(expired)","Show 'Invalid OTP' / 'OTP expired'"],
 ["Login","Email/Username","Email/Text","Required; valid format; must be registered","user@test.com","user@@test","Show 'Invalid email' / 'Account not found'"],
 ["Login","Password","Password","Required; masked input; matched against stored hash","Test@1234","(blank), wrongpass","Show 'Incorrect password' / 'Required field'"],
]
for r in fv_rows:
    ws2.append(r)
last2 = ws2.max_row
style_body(ws2, 2, last2, len(headers2))
for r in range(2, last2 + 1):
    ws2.cell(row=r, column=1).font = Font(name="Arial", bold=True, size=10)
    ws2.row_dimensions[r].height = 32
set_widths(ws2, [16, 20, 14, 34, 20, 20, 34])

# ============================================================
# SHEET 3: BUG REPORT TEMPLATE
# ============================================================
ws3 = wb.create_sheet("Bug Report")
headers3 = ["Bug ID", "Module", "Bug Summary", "Steps to Reproduce", "Test Data Used",
            "Expected Result", "Actual Result", "Severity", "Priority", "Browser/Device",
            "Status", "Reported By", "Date", "Screenshot/Evidence Link"]
ws3.append(headers3)
style_header(ws3)
ws3.freeze_panes = "A2"

# One example row so the format is clear, plus 12 blank rows for actual bugs found
example = ["BUG_001","Sign Up","Signup succeeds even when Confirm Password does not match Password",
           "1. Go to Sign Up\n2. Enter Password: Test@1234\n3. Enter Confirm Password: Test@9999\n4. Click Submit",
           "Password: Test@1234 / Confirm: Test@9999",
           "Form should show 'Passwords do not match' and block submission",
           "Account is created successfully despite mismatched passwords",
           "High","P1","Chrome v127 / Windows 11","Open / New","<Your Name>","2026-09-02","<link/screenshot>"]
ws3.append(example)
for i in range(2, 15):
    ws3.append([f"BUG_{str(i).zfill(3)}"] + [""] * (len(headers3) - 1))

last3 = ws3.max_row
style_body(ws3, 2, last3, len(headers3))
for r in range(2, last3 + 1):
    ws3.row_dimensions[r].height = 40
set_widths(ws3, [10, 14, 30, 32, 20, 26, 26, 10, 10, 16, 12, 14, 12, 20])
# Highlight the example row
for c in range(1, len(headers3)+1):
    ws3.cell(row=2, column=c).fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

# ============================================================
# SHEET 4: README / Instructions
# ============================================================
ws4 = wb.create_sheet("Read Me", 0)
ws4.sheet_view.showGridLines = False
ws4["A1"] = "Manual QA Machine Test — FieldForceConnect"
ws4["A1"].font = Font(name="Arial", bold=True, size=14, color="1F4E78")
notes = [
    "",
    "URL under test: https://test.fieldforceconnect.com/",
    "",
    "How to use this workbook:",
    "1. 'Test Cases' sheet — 47 test cases across Sign Up, Forgot Password, Sign with OTP, and Login modules.",
    "   Fill 'Actual Result' and 'Status' columns as you execute each test manually on the live site.",
    "   If a test fails, log it on the 'Bug Report' sheet and reference the Bug ID in the last column here.",
    "2. 'Field Validations' sheet — expected validation rules for every input field per module,",
    "   with valid/invalid example values to test boundary and negative scenarios.",
    "3. 'Bug Report' sheet — standard bug template (1 filled example row + blank rows to log real bugs found).",
    "",
    "Tip: Update the exact field-level rules (max length, exact regex, OTP digit count, lockout threshold)",
    "once you inspect the live application, since these vary by implementation.",
]
for i, line in enumerate(notes, start=2):
    ws4.cell(row=i, column=1, value=line).font = Font(name="Arial", size=11)
ws4.column_dimensions["A"].width = 110

wb.save("Manual_Test_Cases_FieldForceConnect.xlsx")
print("done")
